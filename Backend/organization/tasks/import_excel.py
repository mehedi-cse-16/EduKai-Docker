import logging
import os
from celery import shared_task

from organization.tasks.geocode import geocode_organization_task

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column name maps — handles variations client might send
# ---------------------------------------------------------------------------
ORG_COLUMN_MAP = {
    "urn":              ["URN", "urn"],
    "name":             ["OrganizationName", "organizationname", "Organization Name"],
    "local_authority":  ["LocalAuthority", "localauthority", "Local Authority"],
    "phase":            ["Phase", "phase"],
    "gender":           ["Gender", "gender"],
    "street":           ["Street", "street"],
    "address_line_1":   ["AddressLine1", "addressline1", "Address Line 1"],
    "address_line_2":   ["AddressLine2", "addressline2", "Address Line 2"],
    "town":             ["Town", "town"],
    "county":           ["County", "county"],
    "postcode":         ["Postcode", "postcode"],
    "telephone":        ["TelephoneNumber", "Telephone", "telephone"],
}

CONTACT_COLUMN_MAP = {
    "name":             ["OrganizationName", "organizationname", "Organization Name"],
    "local_authority":  ["LocalAuthority", "localauthority", "Local Authority"],
    "contact_person":   ["ContactPersonName", "ContactPerson", "contactperson", "Contact Person"],
    "job_title":        ["JobTitle", "Job Title", "jobtitle"],
    "work_email":       ["WorkEmail", "workemail", "Work Email", "Email"],
}

PHASE_MAP = {
    "16 plus":                  "16_plus",
    "16plus":                   "16_plus",
    "all through":              "all_through",
    "middle deemed primary":    "middle_deemed_primary",
    "middle deemed secondary":  "middle_deemed_secondary",
    "not applicable":           "not_applicable",
    "not_applicable":           "not_applicable",
    "nursery":                  "nursery",
    "primary":                  "primary",
    "secondary":                "secondary",
}

GENDER_MAP = {
    "boys":           "boys",
    "boy":            "boys",
    "girls":          "girls",
    "girl":           "girls",
    "mixed":          "mixed",
    "not applicable": "not_applicable",
    "not_applicable": "not_applicable",
}


def _resolve_columns(headers: list, column_map: dict) -> dict:
    """
    Maps model field names to actual column indices in the Excel file.
    Returns dict like: {"name": 2, "local_authority": 0, ...}
    """
    resolved = {}
    for field, possible_names in column_map.items():
        for possible in possible_names:
            for idx, header in enumerate(headers):
                if header and str(header).strip() == possible:
                    resolved[field] = idx
                    break
            if field in resolved:
                break
    return resolved


def _get_cell(row: tuple, col_map: dict, field: str) -> str | None:
    """Safely get a cell value from a row by field name."""
    idx = col_map.get(field)
    if idx is None:
        return None
    val = row[idx]
    if val is None:
        return None
    return str(val).strip() or None


# =============================================================================
# Organization Import Task
# =============================================================================
@shared_task(
    bind=True,
    max_retries=2,
    default_retry_delay=10,
    name="organization.tasks.import_organizations",
)
def import_organizations_task(self, file_path: str):
    """
    Imports organizations from an Excel file.
    Skips rows where name + local_authority already exist in DB.
    Skips rows with missing required fields or bad data.
    Returns a detailed summary.
    """
    import openpyxl
    from organization.models import Organization
    from organization.tasks.geocode import geocode_organization_task

    summary = {
        "total_rows":            0,
        "organizations_created": 0,
        "organizations_skipped": 0,
        "errors":                [],
    }

    try:
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as exc:
            logger.error(f"[import_org] Failed to open file: {exc}")
            raise self.retry(exc=exc)

        if not rows:
            summary["errors"].append("File is empty.")
            return summary

        # ── Map column names to indices ──���────────────────────────────────
        headers = [str(h).strip() if h else "" for h in rows[0]]
        col_map = _resolve_columns(headers, ORG_COLUMN_MAP)

        # Validate required columns exist
        for required in ["name", "local_authority"]:
            if required not in col_map:
                summary["errors"].append(
                    f"Required column missing: '{required}'. "
                    f"Found headers: {headers}"
                )
                return summary

        data_rows = rows[1:]
        summary["total_rows"] = len(data_rows)

        for row_num, row in enumerate(data_rows, start=2):
            # Skip completely empty rows
            if all(cell is None for cell in row):
                summary["total_rows"] -= 1
                continue

            # ── Required fields ───────────────────────────────────────────
            name            = _get_cell(row, col_map, "name")
            local_authority = _get_cell(row, col_map, "local_authority")

            if not name:
                summary["organizations_skipped"] += 1
                summary["errors"].append(f"Row {row_num}: Missing OrganizationName — skipped.")
                continue

            if not local_authority:
                summary["organizations_skipped"] += 1
                summary["errors"].append(f"Row {row_num}: Missing LocalAuthority for '{name}' — skipped.")
                continue

            # ── Skip if already exists ────────────────────────────────────
            if Organization.objects.filter(
                name__iexact=name,
                local_authority__iexact=local_authority,
            ).exists():
                summary["organizations_skipped"] += 1
                logger.debug(f"[import_org] Row {row_num}: '{name}' in '{local_authority}' already exists — skipped.")
                continue

            # ── Parse optional fields ─────────────────────────────────────
            raw_phase  = (_get_cell(row, col_map, "phase") or "").lower().strip()
            phase      = PHASE_MAP.get(raw_phase, "not_applicable")

            raw_gender = (_get_cell(row, col_map, "gender") or "").lower().strip()
            gender     = GENDER_MAP.get(raw_gender, "mixed")

            raw_tel    = _get_cell(row, col_map, "telephone")
            telephone  = str(int(float(raw_tel))) if raw_tel else None

            try:
                org = Organization.objects.create(
                    urn             = _get_cell(row, col_map, "urn"),
                    name            = name,
                    local_authority = local_authority,
                    phase           = phase,
                    gender          = gender,
                    street          = _get_cell(row, col_map, "street"),
                    address_line_1  = _get_cell(row, col_map, "address_line_1"),
                    address_line_2  = _get_cell(row, col_map, "address_line_2"),
                    town            = _get_cell(row, col_map, "town"),
                    county          = _get_cell(row, col_map, "county"),
                    postcode        = _get_cell(row, col_map, "postcode"),
                    telephone       = telephone,
                )
                summary["organizations_created"] += 1

                # ── Queue geocoding for this org ──────────────────────────
                geocode_organization_task.apply_async(
                    args=[str(org.id)],
                    queue="default",
                    countdown=summary["organizations_created"] * 2,
                    # ✅ 2 second stagger — safer than 1s for 24000 orgs
                )

            except Exception as exc:
                summary["organizations_skipped"] += 1
                summary["errors"].append(f"Row {row_num}: Failed to create '{name}' — {exc}")
                logger.error(f"[import_org] Row {row_num}: {exc}")

        logger.info(
            f"[import_org] ✅ Import complete — "
            f"created={summary['organizations_created']}, "
            f"skipped={summary['organizations_skipped']}, "
            f"errors={len(summary['errors'])}"
        )
        from account.utils.activity import log_activity
        log_activity(
            event_type = "org_import_completed",
            severity   = "success" if summary["organizations_created"] > 0 else "warning",
            title      = f"Organizations imported: {summary['organizations_created']}",
            message    = (
                f"Created: {summary['organizations_created']}, "
                f"Skipped: {summary['organizations_skipped']}, "
                f"Errors: {len(summary['errors'])}."
            ),
        )
        return summary

    finally:
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                logger.info(f"[import_org] Temp file deleted: {file_path}")
        except Exception as cleanup_exc:
            logger.warning(f"[import_org] Could not delete temp file: {cleanup_exc}")


# =============================================================================
# Contact Import Task
# =============================================================================
@shared_task(
    bind=True,
    max_retries=2,
    default_retry_delay=10,
    name="organization.tasks.import_contacts",
)
def import_contacts_task(self, file_path: str):
    """
    Imports contacts from an Excel file.
    Matches contacts to organizations via OrganizationName + LocalAuthority.
    Skips rows where work_email already exists in DB.
    Skips rows with missing required fields or invalid email.
    Returns a detailed summary.
    """
    import re
    import openpyxl
    from organization.models import Organization, OrganizationContact

    EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

    summary = {
        "total_rows":       0,
        "contacts_created": 0,
        "contacts_skipped": 0,
        "errors":           [],
    }

    try:
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as exc:
            logger.error(f"[import_contact] Failed to open file: {exc}")
            raise self.retry(exc=exc)

        if not rows:
            summary["errors"].append("File is empty.")
            return summary

        # ── Map column names to indices ───────────────────────────────────
        headers = [str(h).strip() if h else "" for h in rows[0]]
        col_map = _resolve_columns(headers, CONTACT_COLUMN_MAP)

        # Validate required columns
        for required in ["name", "local_authority", "contact_person", "work_email"]:
            if required not in col_map:
                summary["errors"].append(
                    f"Required column missing: '{required}'. "
                    f"Found headers: {headers}"
                )
                return summary

        data_rows = rows[1:]
        summary["total_rows"] = len(data_rows)

        # ── Cache org lookups to avoid repeated DB hits ───────────────────
        org_cache = {}

        for row_num, row in enumerate(data_rows, start=2):
            if all(cell is None for cell in row):
                summary["total_rows"] -= 1
                continue

            # ── Required fields ───────────────────────────────────────────
            org_name        = _get_cell(row, col_map, "name")
            local_authority = _get_cell(row, col_map, "local_authority")
            contact_person  = _get_cell(row, col_map, "contact_person")
            work_email      = _get_cell(row, col_map, "work_email")

            if not org_name:
                summary["contacts_skipped"] += 1
                summary["errors"].append(f"Row {row_num}: Missing OrganizationName — skipped.")
                continue

            if not local_authority:
                summary["contacts_skipped"] += 1
                summary["errors"].append(f"Row {row_num}: Missing LocalAuthority — skipped.")
                continue

            if not contact_person:
                summary["contacts_skipped"] += 1
                summary["errors"].append(f"Row {row_num}: Missing ContactPerson for '{org_name}' — skipped.")
                continue

            if not work_email:
                summary["contacts_skipped"] += 1
                summary["errors"].append(f"Row {row_num}: Missing WorkEmail for '{contact_person}' — skipped.")
                continue

            # ── Validate email format ─────────────────────────────────────
            work_email = work_email.lower().strip()
            if not EMAIL_REGEX.match(work_email):
                summary["contacts_skipped"] += 1
                summary["errors"].append(
                    f"Row {row_num}: Invalid email '{work_email}' "
                    f"for '{contact_person}' — skipped."
                )
                continue

            # ── Skip if email already exists ─────────────────────────────
            if OrganizationContact.objects.filter(work_email=work_email).exists():
                summary["contacts_skipped"] += 1
                logger.debug(f"[import_contact] Row {row_num}: '{work_email}' already exists — skipped.")
                continue

            # ── Find organization ────────────────────────────────────────
            cache_key = f"{org_name.lower()}::{local_authority.lower()}"
            if cache_key not in org_cache:
                org = Organization.objects.filter(
                    name__iexact=org_name,
                    local_authority__iexact=local_authority,
                ).first()
                org_cache[cache_key] = org

            org = org_cache[cache_key]

            if not org:
                summary["contacts_skipped"] += 1
                summary["errors"].append(
                    f"Row {row_num}: Organization '{org_name}' with "
                    f"LocalAuthority='{local_authority}' not found in DB — skipped. "
                    f"Check if LocalAuthority matches exactly with the organization file."
                )
                continue

            # ── Create contact ───────────────────────────────────────────
            try:
                OrganizationContact.objects.create(
                    organization   = org,
                    contact_person = contact_person,
                    job_title      = _get_cell(row, col_map, "job_title"),
                    work_email     = work_email,
                )
                summary["contacts_created"] += 1

            except Exception as exc:
                summary["contacts_skipped"] += 1
                summary["errors"].append(
                    f"Row {row_num}: Failed to create contact "
                    f"'{contact_person}' — {exc}"
                )
                logger.error(f"[import_contact] Row {row_num}: {exc}")

        logger.info(
            f"[import_contact] ✅ Import complete — "
            f"created={summary['contacts_created']}, "
            f"skipped={summary['contacts_skipped']}, "
            f"errors={len(summary['errors'])}"
        )
        from account.utils.activity import log_activity
        log_activity(
            event_type = "contact_import_completed",
            severity   = "success" if summary["contacts_created"] > 0 else "warning",
            title      = f"Contacts imported: {summary['contacts_created']}",
            message    = (
                f"Created: {summary['contacts_created']}, "
                f"Skipped: {summary['contacts_skipped']}, "
                f"Errors: {len(summary['errors'])}."
            ),
        )
        return summary

    finally:
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                logger.info(f"[import_contact] Temp file deleted: {file_path}")
        except Exception as cleanup_exc:
            logger.warning(f"[import_contact] Could not delete temp file: {cleanup_exc}")